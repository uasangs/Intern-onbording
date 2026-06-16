from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.models.models import InternRecord
from datetime import date
import io


def generate_fy_tracker_excel(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "FY 2025-26"

    # Colours matching Grasim Excel
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    TITLE_FILL = PatternFill("solid", fgColor="2E75B6")
    TITLE_FONT = Font(color="FFFFFF", bold=True, size=12)
    BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    ALT_FILL = PatternFill("solid", fgColor="DEEAF1")

    # Title row
    ws.merge_cells("A1:AJ1")
    ws["A1"] = "FY - 2025-26 | Intern Tracker | Grasim Industries Ltd. (MBDD / TRADC)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Header columns (matching original FY tracker columns)
    static_headers = [
        "Sr. No.", "Name", "Source", "Location", "Name of the Institute",
        "Qualification", "Department", "Project Manager",
        "Start Date", "End Date", "Stipend", "Mobile", "Contact No.",
        "Bank Name", "Account No.", "IFSC Code",
        "Feedback", "Project Submission", "Project Presentation (Y/N)",
        "Evaluation from Project Mgr (Y/N)", "Panel Evaluation",
        "Experience Certificate given (Y/N)", "ABG Email ID",
    ]

    # Monthly stipend columns (Apr → Mar)
    months = [
        "Apr 25", "May 25", "Jun 25", "Jul 25", "Aug 25", "Sep 25",
        "Oct 25", "Nov 25", "Dec 25", "Jan 26", "Feb 26", "Mar 26"
    ]
    all_headers = static_headers + months

    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[2].height = 30

    # Data rows
    records = db.query(InternRecord).order_by(InternRecord.serial_no).all()

    for row_idx, record in enumerate(records, start=3):
        c = record.candidate
        bank = c.bank_details if c else None
        it = record.it_task
        review = record.manager_review
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()

        def _y_n(val: bool) -> str:
            return "Y" if val else "N"

        # Build stipend payment map
        stipend_map = {}
        if record.accounts_task:
            for payment in record.accounts_task.stipend_payments:
                if payment.month_year:
                    stipend_map[payment.month_year] = (
                        f"₹{int(payment.amount):,}" if payment.status == "paid" else "Pending"
                    )

        row_data = [
            record.serial_no or row_idx - 2,
            c.full_name if c else "",
            record.source or "",
            record.location or "",
            c.institute_name if c else "",
            c.qualification if c else "",
            record.department or "",
            record.reporting_manager.name if record.reporting_manager else "",
            record.start_date.strftime("%d/%m/%Y") if record.start_date else "",
            record.end_date.strftime("%d/%m/%Y") if record.end_date else "",
            f"₹{int(record.stipend_amount):,}" if record.stipend_amount else "",
            c.mobile if c else "",
            c.contact_no if c else "",
            bank.bank_name if bank else "",
            bank.account_number if bank else "",
            bank.ifsc_code if bank else "",
            "",  # Feedback (from form)
            _y_n(record.project_submission_done),
            _y_n(record.project_presentation_done),
            _y_n(record.eval_from_mgr_done),
            _y_n(record.panel_evaluation_done),
            _y_n(record.experience_certificate_issued),
            it.abg_email_id if it else "",
        ] + [stipend_map.get(m, "") for m in months]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            if fill.fill_type:
                cell.fill = fill

    # Column widths
    col_widths = [6, 22, 12, 10, 28, 16, 16, 20, 12, 12, 10, 14, 14, 18, 20, 14, 10, 16, 18, 20, 14, 20, 26]
    col_widths += [10] * 12  # monthly columns

    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header rows
    ws.freeze_panes = "A3"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
